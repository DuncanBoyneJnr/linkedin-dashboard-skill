#!/usr/bin/env python3
"""
Step 5a: Extract and merge LinkedIn Analytics XLSX exports into persistent archive.

Scans the project directory for all .xlsx and .xls files, auto-detects format
by sheet names (no manual format selection needed), merges into
analytics_archive.json, and builds analytics_full.json with derived datasets.

Also handles .zip archives — extracts them into a temp directory and processes
any .xlsx/.xls files found inside.
"""

import json
import os
import re
import shutil
import sys
import tempfile
import zipfile
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

# Accept project directory as CLI arg, default to current dir
if len(sys.argv) > 1:
    PROJECT_DIR = Path(sys.argv[1]).resolve()
else:
    PROJECT_DIR = Path.cwd()

ARCHIVE_PATH = PROJECT_DIR / "analytics_archive.json"
DATA_PATH = PROJECT_DIR / "analytics_data.json"
FULL_PATH = PROJECT_DIR / "analytics_full.json"


# ── xlrd fallback ──────────────────────────────────────────────────────────

def load_workbook_safe(filepath, read_only=False):
    """
    Open an XLSX/XLS file with openpyxl, falling back to xlrd for old binary
    .xls files (CDFV2 format) that openpyxl cannot read.
    """
    from openpyxl import load_workbook
    try:
        return load_workbook(filename=filepath, read_only=read_only)
    except (zipfile.BadZipFile, Exception) as e:
        # Try xlrd for old binary .xls files
        try:
            import xlrd
        except ImportError:
            raise RuntimeError(
                "xlrd is required for old binary .xls files. "
                "Install with: pip3 install xlrd"
            ) from e

        xl_book = xlrd.open_workbook(str(filepath))
        return XlrdWorkbookWrapper(xl_book)


class XlrdSheetWrapper:
    """Wraps an xlrd sheet to match openpyxl's worksheet interface."""

    def __init__(self, xl_sheet):
        self._sheet = xl_sheet
        self.title = xl_sheet.name

    def iter_rows(self, min_row=1, max_col=None, values_only=False):
        """Yield rows matching openpyxl's iter_rows(values_only=True)."""
        ncols = self._sheet.ncols
        if max_col is not None:
            ncols = min(ncols, max_col)
        for r in range(min_row - 1, self._sheet.nrows):
            row = []
            for c in range(ncols):
                cell = self._sheet.cell(r, c)
                row.append(cell.value)
            yield row


class XlrdWorkbookWrapper:
    """Wraps an xlrd workbook to match openpyxl's workbook interface."""

    def __init__(self, xl_book):
        self._book = xl_book
        self.worksheets = [XlrdSheetWrapper(s) for s in xl_book.sheets()]

    def __getitem__(self, name):
        for ws in self.worksheets:
            if ws.title == name:
                return ws
        raise KeyError(f"Sheet '{name}' not found")

    def close(self):
        pass


# ── Helper functions ──────────────────────────────────────────────────────

def safe_float(val):
    """Coerce a cell value to float, handling text strings and percentage strings."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return 0.0
    # Handle "< 1%" or "2%" style strings
    is_pct = "%" in s
    s = s.replace("<", "").replace("%", "").strip()
    try:
        result = float(s)
        # If the original was a percentage string (e.g. "23%"), divide by 100
        # If it was a plain number string (e.g. "1" meaning 1%), also divide by 100
        if is_pct or (not isinstance(val, (int, float)) and result <= 100):
            result = result / 100.0
        return result
    except ValueError:
        return 0.0


def safe_int(val):
    """Coerce a cell value to int, handling text strings."""
    if val is None:
        return 0
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    s = str(val).strip()
    if not s:
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def parse_date(date_string):
    """Parse a date string to YYYY-MM-DD format."""
    from dateutil.parser import parse
    return parse(str(date_string)).strftime("%Y-%m-%d")


# ── Format A: Personal/Creator Analytics (AggregateAnalytics_* / Content_*) ──

def extract_engagement(sheet):
    """Extract daily engagement data from ENGAGEMENT sheet."""
    rows = []
    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        if row[0] is None:
            continue
        try:
            date = parse_date(row[0])
            rows.append({
                "date": date,
                "impressions": safe_int(row[1]),
                "engagements": safe_int(row[2]),
            })
        except Exception:
            continue
    return rows


def extract_followers(sheet):
    """Extract daily follower data from FOLLOWERS sheet."""
    rows = []
    total_snapshots = []
    for row in sheet.iter_rows(min_row=4, max_col=3, values_only=True):
        if row[0] is None:
            continue
        val = str(row[0]).strip()
        # Check for total snapshot like "Total followers on 6/5/2026"
        if "total followers" in val.lower():
            # Extract date with regex
            m = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", val)
            if m:
                total_snapshots.append({
                    "date": parse_date(m.group(1)),
                    "total": safe_int(row[1]),
                })
            continue
        try:
            date = parse_date(row[0])
            rows.append({
                "date": date,
                "new_followers": safe_int(row[1]),
            })
        except Exception:
            continue
    return rows, total_snapshots


def extract_demographics(sheet, start_date, end_date):
    """Extract demographics data from DEMOGRAPHICS sheet."""
    rows = []
    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        pct = safe_float(row[2])
        rows.append({
            "demographic_type": str(row[0]).strip(),
            "value": str(row[1]).strip(),
            "percentage": pct,
            "start_date": start_date,
            "end_date": end_date,
        })
    return rows


def extract_top_posts(sheet):
    """Extract top posts from TOP POSTS sheet (two tables)."""
    posts = {}

    # Table 1: cols A-C (by engagements) - min_row=4
    for row in sheet.iter_rows(min_row=4, min_col=1, max_col=3, values_only=True):
        if row[0] is None:
            continue
        url = str(row[0]).strip()
        if url:
            posts[url] = {
                "post_url": url,
                "publish_date": parse_date(row[1]) if row[1] else "",
                "impressions": 0,
                "engagements": safe_int(row[2]),
            }

    # Table 2: cols E-G (by impressions) - min_row=4
    for row in sheet.iter_rows(min_row=4, min_col=5, max_col=7, values_only=True):
        if row[0] is None:
            continue
        url = str(row[0]).strip()
        if url in posts:
            posts[url]["impressions"] = safe_int(row[2])
            posts[url]["publish_date"] = parse_date(row[1]) if row[1] else posts[url]["publish_date"]
        else:
            posts[url] = {
                "post_url": url,
                "publish_date": parse_date(row[1]) if row[1] else "",
                "impressions": safe_int(row[2]),
                "engagements": 0,
            }

    return list(posts.values())


# ── Format B: Competitor Analytics (single COMPETITORS sheet) ──────────────

def extract_competitor_analytics(filepath):
    """Extract competitor analytics data from a COMPETITORS sheet.

    Format:
      Row 1: Date range (start date in col A, end date in col B)
      Row 2: Headers (Page, New Followers, Posts, Comments, Comments per day, Reactions)
      Row 3+: Data rows for each company page
    """
    wb = load_workbook_safe(filepath)  # NOT read_only — may miss rows
    sheet = wb["COMPETITORS"]

    rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 3:
        return None

    # Row 1: date range
    date_range = {}
    if rows[0][0]:
        date_range["start"] = parse_date(rows[0][0])
    if rows[0][1]:
        date_range["end"] = parse_date(rows[0][1])

    # Row 2: headers
    headers = [str(h).strip() if h else "" for h in rows[1]]

    # Row 3+: data
    data_rows = []
    for row in rows[2:]:
        if row[0] is None:
            continue
        entry = {}
        for i, h in enumerate(headers):
            if i < len(row):
                val = row[i]
                # Convert numeric fields
                if h in ("New Followers", "Posts", "Comments", "Comments per day", "Reactions"):
                    entry[h] = safe_int(val)
                else:
                    entry[h] = str(val).strip() if val else ""
        data_rows.append(entry)

    return {
        "source_type": "competitor",
        "date_range": date_range,
        "headers": headers,
        "rows": data_rows,
    }


# ── Format C: Company Page Analytics (New followers / Metrics / All posts / Visitor metrics) ──

def extract_followers_new(sheet):
    """Extract daily follower data from 'New followers' sheet (company page format)."""
    rows = []
    for row in sheet.iter_rows(min_row=2, max_col=5, values_only=True):
        if row[0] is None:
            continue
        try:
            date = parse_date(row[0])
            rows.append({
                "date": date,
                "sponsored_followers": safe_float(row[1]) if len(row) > 1 else 0.0,
                "organic_followers": safe_float(row[2]) if len(row) > 2 else 0.0,
                "auto_invited_followers": safe_float(row[3]) if len(row) > 3 else 0.0,
                "new_followers": safe_float(row[4]) if len(row) > 4 else 0.0,
            })
        except Exception:
            continue
    return rows


def extract_engagement_new(sheet):
    """Extract daily engagement data from 'Metrics' sheet (company page format)."""
    rows = []
    for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
        if row[0] is None:
            continue
        try:
            date = parse_date(row[0])
            rows.append({
                "date": date,
                "impressions": safe_float(row[1]) if len(row) > 1 else 0.0,
                "unique_impressions": safe_float(row[2]) if len(row) > 2 else 0.0,
                "engagement_rate": safe_float(row[3]) if len(row) > 3 else 0.0,
            })
        except Exception:
            continue
    return rows


def extract_all_posts_new(sheet):
    """Extract per-post data from 'All posts' sheet (company page format)."""
    posts = []
    for row in sheet.iter_rows(min_row=2, max_col=5, values_only=True):
        if row[0] is None:
            continue
        try:
            posts.append({
                "date": parse_date(row[0]),
                "post_url": str(row[1]).strip() if row[1] else "",
                "impressions": safe_float(row[2]) if len(row) > 2 else 0.0,
                "engagements": safe_float(row[3]) if len(row) > 3 else 0.0,
                "engagement_rate": safe_float(row[4]) if len(row) > 4 else 0.0,
            })
        except Exception:
            continue
    return posts


def extract_visitors_new(sheet):
    """Extract daily visitor data from 'Visitor metrics' sheet (company page format)."""
    rows = []
    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        if row[0] is None:
            continue
        try:
            rows.append({
                "date": parse_date(row[0]),
                "page_views": safe_float(row[1]) if len(row) > 1 else 0.0,
                "unique_visitors": safe_float(row[2]) if len(row) > 2 else 0.0,
            })
        except Exception:
            continue
    return rows


def extract_demographics_new(wb, sheet_names):
    """Extract demographics from company page format sheets (Location, Job function, etc.)."""
    demographic_sheets = [
        "Location", "Job function", "Seniority", "Industry", "Company size"
    ]
    demographics = []
    for name in demographic_sheets:
        if name in sheet_names:
            sheet = wb[name]
            for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[0] is None:
                    continue
                demographics.append({
                    "demographic_type": name,
                    "value": str(row[0]).strip(),
                    "total_followers": safe_float(row[1]) if len(row) > 1 else 0.0,
                })
    return demographics


# ── Main extraction dispatcher ─────────────────────────────────────────────

def extract_xlsx(filepath):
    """Extract all data from a single XLSX/XLS file, auto-detecting format by sheet names."""
    wb = load_workbook_safe(filepath, read_only=True)
    result = {}

    sheet_names = [s.title for s in wb.worksheets]

    # Detect competitor analytics (single COMPETITORS sheet)
    if sheet_names == ["COMPETITORS"]:
        wb.close()
        comp = extract_competitor_analytics(filepath)
        if comp:
            result["competitor"] = comp
        return result

    # Detect company page format (New followers / Metrics / All posts / Visitor metrics)
    if "New followers" in sheet_names or "Metrics" in sheet_names:
        if "New followers" in sheet_names:
            result["followers"] = extract_followers_new(wb["New followers"])
            result["follower_demographics"] = extract_demographics_new(wb, sheet_names)
        if "Metrics" in sheet_names:
            result["engagement"] = extract_engagement_new(wb["Metrics"])
        if "All posts" in sheet_names:
            result["top_posts"] = extract_all_posts_new(wb["All posts"])
        if "Visitor metrics" in sheet_names:
            result["visitors"] = extract_visitors_new(wb["Visitor metrics"])
            result["visitor_demographics"] = extract_demographics_new(wb, sheet_names)
        wb.close()
        return result

    # Detect personal/creator format (ENGAGEMENT / FOLLOWERS / DEMOGRAPHICS / TOP POSTS)
    if "ENGAGEMENT" in sheet_names:
        result["engagement"] = extract_engagement(wb["ENGAGEMENT"])
    if "FOLLOWERS" in sheet_names:
        followers, snapshots = extract_followers(wb["FOLLOWERS"])
        result["followers"] = followers
        result["follower_snapshots"] = snapshots
    if "DEMOGRAPHICS" in sheet_names:
        # Extract date range from filename
        fname = Path(filepath).name
        dates = fname.split("_")
        start_date = ""
        end_date = ""
        for i, part in enumerate(dates):
            if re.match(r"\d{4}-\d{2}-\d{2}", part):
                if not start_date:
                    start_date = part
                else:
                    end_date = part
        result["demographics"] = extract_demographics(wb["DEMOGRAPHICS"], start_date, end_date)
    if "TOP POSTS" in sheet_names:
        result["top_posts"] = extract_top_posts(wb["TOP POSTS"])

    wb.close()
    return result


# ── Archive merge ──────────────────────────────────────────────────────────

def merge_archives(existing, new_data, filepath):
    """Merge new data into existing archive, keeping higher values on overlap."""
    fname = Path(filepath).name

    # Engagement
    existing_eng = existing.get("engagement", {})
    for row in new_data.get("engagement", []):
        d = row["date"]
        if d in existing_eng:
            existing_eng[d]["impressions"] = max(existing_eng[d]["impressions"], row["impressions"])
            existing_eng[d]["engagements"] = max(existing_eng[d]["engagements"], row["engagements"])
        else:
            existing_eng[d] = row
    existing["engagement"] = existing_eng

    # Followers
    existing_fol = existing.get("followers", {})
    for row in new_data.get("followers", []):
        d = row["date"]
        if d in existing_fol:
            existing_fol[d]["new_followers"] = max(existing_fol[d]["new_followers"], row["new_followers"])
        else:
            existing_fol[d] = row
    existing["followers"] = existing_fol

    # Follower snapshots
    existing_snaps = existing.get("follower_snapshots", [])
    existing_snaps.extend(new_data.get("follower_snapshots", []))
    existing["follower_snapshots"] = existing_snaps

    # Top posts
    existing_posts = existing.get("top_posts", {})
    for post in new_data.get("top_posts", []):
        url = post["post_url"]
        if url in existing_posts:
            existing_posts[url]["impressions"] = max(existing_posts[url]["impressions"], post["impressions"])
            existing_posts[url]["engagements"] = max(existing_posts[url]["engagements"], post["engagements"])
        else:
            existing_posts[url] = post
    existing["top_posts"] = existing_posts

    # Demographics - use the file with widest date range
    if new_data.get("demographics"):
        existing["demographics"] = new_data["demographics"]

    return existing


# ── Derived datasets ───────────────────────────────────────────────────────

def build_derived(archive):
    """Build derived datasets: monthly aggregates, cumulative followers, totals."""
    daily = []
    # Build daily from engagement + followers
    all_dates = set()
    all_dates.update(archive.get("engagement", {}).keys())
    all_dates.update(archive.get("followers", {}).keys())

    for date_str in sorted(all_dates):
        entry = {"date": date_str}
        eng = archive.get("engagement", {}).get(date_str, {})
        entry["impressions"] = eng.get("impressions", 0)
        entry["engagements"] = eng.get("engagements", 0)
        fol = archive.get("followers", {}).get(date_str, {})
        entry["new_followers"] = fol.get("new_followers", 0)
        daily.append(entry)

    # Monthly aggregates
    monthly = defaultdict(lambda: {
        "impressions": 0, "engagements": 0, "new_followers": 0, "post_count": 0
    })
    for entry in daily:
        month = entry["date"][:7]
        monthly[month]["impressions"] += entry["impressions"]
        monthly[month]["engagements"] += entry["engagements"]
        monthly[month]["new_followers"] += entry["new_followers"]

    monthly_list = []
    for month in sorted(monthly.keys()):
        m = monthly[month]
        eng_rate = m["engagements"] / m["impressions"] if m["impressions"] > 0 else 0
        monthly_list.append({
            "month": month,
            "impressions": m["impressions"],
            "engagements": m["engagements"],
            "new_followers": m["new_followers"],
            "engagement_rate": round(eng_rate, 4),
        })

    # Cumulative followers
    snapshots = archive.get("follower_snapshots", [])
    followers_daily = archive.get("followers", {})
    sorted_dates = sorted(followers_daily.keys())

    cumulative = []
    if snapshots:
        # Use earliest snapshot as anchor
        earliest_snap = min(snapshots, key=lambda s: s["date"])
        anchor_date = earliest_snap["date"]
        anchor_total = earliest_snap["total"]

        # Reconstruct backwards and forwards
        cum_map = {}
        cum_map[anchor_date] = anchor_total

        # Forward
        running = anchor_total
        for d in sorted_dates:
            if d > anchor_date:
                running += followers_daily[d].get("new_followers", 0)
                cum_map[d] = running

        # Backward
        running = anchor_total
        for d in reversed(sorted_dates):
            if d < anchor_date:
                running -= followers_daily[d].get("new_followers", 0)
                cum_map[d] = running

        cumulative = [{"date": d, "total": cum_map[d]} for d in sorted(cum_map.keys())]

    # Totals
    totals = {
        "total_impressions": sum(e["impressions"] for e in daily),
        "total_engagements": sum(e["engagements"] for e in daily),
        "total_new_followers": sum(e["new_followers"] for e in daily),
        "total_days": len(daily),
        "avg_daily_impressions": round(sum(e["impressions"] for e in daily) / len(daily), 1) if daily else 0,
        "avg_engagement_rate": round(
            sum(e["engagements"] for e in daily) / sum(e["impressions"] for e in daily), 4
        ) if daily and sum(e["impressions"] for e in daily) > 0 else 0,
    }

    result = {
        "daily": daily,
        "monthly": monthly_list,
        "top_posts": list(archive.get("top_posts", {}).values()),
        "demographics": archive.get("demographics", []),
        "cumulative_followers": cumulative,
        "totals": totals,
    }

    if archive.get("competitor"):
        result["competitor"] = archive["competitor"]

    return result


# ── File discovery ─────────────────────────────────────────────────────────

def find_spreadsheet_files(project_dir):
    """
    Find all .xlsx and .xls files in the project directory.
    Also checks for .zip archives and extracts them into a temp directory.
    Returns a list of (filepath, is_temp) tuples.
    """
    files = []

    # Scan for .xlsx and .xls files
    for ext in ("*.xlsx", "*.xls"):
        files.extend(project_dir.glob(ext))

    # Check for .zip archives and extract them
    for zip_path in project_dir.glob("*.zip"):
        print(f"\nFound zip archive: {zip_path.name}")
        try:
            with zipfile.ZipFile(zip_path, 'r') as zf:
                # Extract to a temp directory
                extract_dir = Path(tempfile.mkdtemp(prefix="linkedin_zip_"))
                zf.extractall(path=extract_dir)
                print(f"  Extracted to {extract_dir}")
                for ext in ("*.xlsx", "*.xls"):
                    files.extend(extract_dir.glob(ext))
        except Exception as e:
            print(f"  WARNING: Could not extract {zip_path.name}: {e}")

    # Deduplicate by resolved path
    seen = set()
    unique = []
    for f in files:
        resolved = f.resolve()
        if resolved not in seen:
            seen.add(resolved)
            unique.append(resolved)

    return sorted(unique)


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    # Find all spreadsheet files (auto-detect, no hardcoded filename patterns)
    xlsx_files = find_spreadsheet_files(PROJECT_DIR)

    if not xlsx_files:
        print("No LinkedIn Analytics XLSX/XLS files found.")
        print(f"Looked in: {PROJECT_DIR}")
        print("Drop your LinkedIn Analytics exports (.xlsx or .xls) or a .zip archive into this folder.")
        return

    print(f"Found {len(xlsx_files)} spreadsheet file(s):")
    for f in xlsx_files:
        print(f"  - {f.name}")

    # Load existing archive if present
    archive = {}
    if ARCHIVE_PATH.exists():
        with open(ARCHIVE_PATH) as f:
            archive = json.load(f)
        # Backup
        backup_path = PROJECT_DIR / f"analytics_archive.backup-{datetime.now().strftime('%Y%m%d')}.json"
        shutil.copy2(ARCHIVE_PATH, backup_path)
        print(f"\nBacked up existing archive to {backup_path.name}")
        print(f"Archive had {len(archive.get('engagement', {}))} engagement days")

    # Process each file
    for fpath in xlsx_files:
        print(f"\nProcessing {fpath.name}...")
        try:
            data = extract_xlsx(fpath)
            if "competitor" in data:
                comp = data["competitor"]
                print(f"  Competitor analytics: {len(comp['rows'])} companies, "
                      f"date range: {comp['date_range'].get('start', '?')} - {comp['date_range'].get('end', '?')}")
                archive["competitor"] = comp
            else:
                eng_count = len(data.get("engagement", []))
                fol_count = len(data.get("followers", []))
                post_count = len(data.get("top_posts", []))
                demo_count = len(data.get("demographics", []))
                print(f"  Engagement: {eng_count} days, Followers: {fol_count} days, "
                      f"Top Posts: {post_count}, Demographics: {demo_count}")
                archive = merge_archives(archive, data, fpath)
        except Exception as e:
            print(f"  ERROR: {e}")

    # Write archive
    with open(ARCHIVE_PATH, "w") as f:
        json.dump(archive, f, indent=2, default=str)
    print(f"\nWrote analytics_archive.json")

    # Write working copy
    with open(DATA_PATH, "w") as f:
        json.dump(archive, f, indent=2, default=str)
    print(f"Wrote analytics_data.json")

    # Build and write derived datasets
    derived = build_derived(archive)
    with open(FULL_PATH, "w") as f:
        json.dump(derived, f, indent=2, default=str)
    print(f"Wrote analytics_full.json")

    # Summary
    print(f"\n--- Summary ---")
    print(f"Engagement days: {len(archive.get('engagement', {}))}")
    print(f"Follower days: {len(archive.get('followers', {}))}")
    print(f"Top posts: {len(archive.get('top_posts', {}))}")
    print(f"Demographics: {len(archive.get('demographics', []))}")
    print(f"Follower snapshots: {len(archive.get('follower_snapshots', []))}")
    print(f"Monthly aggregates: {len(derived['monthly'])} months")
    print(f"Total impressions: {derived['totals']['total_impressions']:,}")
    print(f"Total engagements: {derived['totals']['total_engagements']:,}")
    if archive.get("competitor"):
        comp = archive["competitor"]
        print(f"\nCompetitor analytics: {len(comp['rows'])} companies")
        for r in comp["rows"]:
            print(f"  - {r.get('Page', '?'):30s} | Followers: {r.get('New Followers', 0):>4d} | "
                  f"Posts: {r.get('Posts', 0):>3d} | Reactions: {r.get('Reactions', 0):>3d}")


if __name__ == "__main__":
    main()
